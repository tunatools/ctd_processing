import pathlib
import openpyxl
import pandas as pd
import shutil


def get_data_from_column_file(file_path):
    data = {}
    header = []
    with open(file_path) as fid:
        for r, line in enumerate(fid):
            if not line.strip():
                continue
            split_line = [item.strip() for item in line.split('\t')]
            if r == 0:
                data = dict((key, []) for key in split_line)
                header = split_line
            else:
                for key, value in zip(header, split_line):
                    data[key].append(value)
    return data


class DeliveryMetadataFile:
    def __init__(self):
        self._sensorinfo = {}
        self._metadata = {}

        self._sensorinfo_columns = []
        self._metadata_columns = []

        self._sensorinfo_sheet_name = 'Sensorinfo'
        self._metadata_sheet_name = 'Metadata'

        self._sensorinfo_header_row = 2
        self._metadata_header_row = 2

        self.template_path = pathlib.Path(pathlib.Path(__file__).parent, 'resources', 'metadata_template.xlsx')

        self.wb = openpyxl.load_workbook(self.template_path)

        self._load_headers()

    def _load_headers(self):
        self._sensorinfo_columns = self._get_header_in_worksheet(self._sensorinfo_sheet_name, self._sensorinfo_header_row)
        self._metadata_columns = self._get_header_in_worksheet(self._metadata_sheet_name, self._metadata_header_row)

    def _get_header_in_worksheet(self, sheet_name, row_nr):
        ws = self.wb[sheet_name]
        for r, row in enumerate(ws.iter_rows()):
            if r == row_nr:
                return [cell.value for cell in row]

    def add_sensorinfo_from_file(self, file_path):
        self._sensorinfo = get_data_from_column_file(file_path)

    def add_metadata_from_file(self, file_path):
        self._metadata = get_data_from_column_file(file_path)

    def add_delivery_note_from_file(self, file_path):
        pass

    def old_save_file(self, save_path):
        self._write_sensorinfo()
        self._write_metadata()
        self.wb.save(save_path)

    def save_file(self, save_path):
        self._copy_template(save_path)
        self._read_sheets(save_path)
        self._write_sensorinfo(save_path)
        self._write_metadata(save_path)

    def _copy_template(self, save_path):
        shutil.copy2(self.template_path, save_path)

    def _write_sensorinfo(self, file_path):
        d = self._sensorinfo.copy()
        for col in ['MET_COMNT.1', 'TIME']:
            if col in d:
                print(True)
                d.pop(col)
        dd = {'Tabellhuvud:': ['' for len in d['INSTRUMENT_SERIE']]}
        dd.update(d)
        df = pd.DataFrame(dd)

        tdf = pd.read_excel(file_path, sheet_name='Sensorinfo')
        tdf.columns = tdf.iloc[1]

        new_df = pd.concat([tdf, df])
        self._write_df_to_xlsx(new_df, file_path, 'Sensorinfo')

    def _write_metadata(self, file_path):
        d = self._metadata.copy()
        for col in ['MET_COMNT.1', 'TIME']:
            if col in d:
                print(True)
                d.pop(col)
        dd = {'Tabellhuvud:': ['' for len in d['MYEAR']]}
        dd.update(d)
        df = pd.DataFrame(dd)

        tdf = pd.read_excel(file_path, sheet_name='Metadata')
        tdf.columns = tdf.iloc[1]

        new_df = pd.concat([tdf, df])
        self._write_df_to_xlsx(new_df, file_path, 'Metadata')

    def _write_df_to_xlsx(self, df, file_path, sheet_name):
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
            wb = writer.book
            try:
                wb.remove(wb[sheet_name])
            except:
                print("Worksheet does not exist")
            finally:
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=['' for _ in df.columns], encoding='cp1252')
                writer.save()

    def old_write_sensorinfo(self):
        if not self._sensorinfo:
            return
        ws = self.wb[self._sensorinfo_sheet_name]
        nr_rows = len(list(self._sensorinfo.values())[0])
        print(nr_rows)
        for c, col in enumerate(self._sensorinfo_columns):
            for i in range(nr_rows):
                r = i + self._sensorinfo_header_row + 2
                value_list = self._sensorinfo.get(col)
                if not value_list:
                    value = ''
                else:
                    value = self._sensorinfo[col][i]
                ws.cell(r, c+1).value = value

    def old_write_metadata(self):
        if not self._metadata:
            return
        ws = self.wb[self._metadata_sheet_name]
        nr_rows = len(list(self._metadata.values())[0])
        print(nr_rows)
        for c, col in enumerate(self._metadata_columns):
            for i in range(nr_rows):
                r = i + self._metadata_header_row + 2
                value_list = self._metadata.get(col)
                if not value_list:
                    value = ''
                else:
                    value = self._metadata[col][i]
                ws.cell(r, c+1).value = value


if __name__ == '__main__':
    sensor = r'C:\mw\temp_svea_local\temp\delivery_files\20211125142926/sensorinfo.txt'
    d = get_data_from_column_file(sensor)
    for col in ['MET_COMNT.1', 'TIME']:
        if col in d:
            print(True)
            d.pop(col)
    dd = {'Tabellhuvud:': ['' for len in d['INSTRUMENT_SERIE']]}
    dd.update(d)
    df = pd.DataFrame(dd)

    template_path = pathlib.Path(pathlib.Path(__file__).parent, 'resources', 'metadata_template.xlsx')
    template_path_out = pathlib.Path(template_path.parent, 'metadata_test.xlsx')

    xlxs = pd.ExcelFile(template_path_out)
    sheets = {}
    sheet_name = ['Förklaring', 'Metadata', 'Sensorinfo', 'Information']
    header_row = [None, 2, 2, 0]
    for sheet, h_row in zip(sheet_name, header_row):
        sheets[sheet] = xlxs.parse(
            sheet,
            header=h_row,
            dtype=str,
            keep_default_na=False
        )


    if 0:
        import shutil
        from openpyxl import load_workbook
        shutil.copy2(template_path, template_path_out)

        tdf = pd.read_excel(template_path, sheet_name='Sensorinfo')
        tdf.columns = tdf.iloc[1]

        new_df = pd.concat([tdf, df])

        with pd.ExcelWriter(template_path_out, engine='openpyxl', mode='a') as writer:
            wb = writer.book
            try:
                wb.remove(wb['Sensorinfo'])
            except:
                print("Worksheet does not exist")
            finally:
                new_df.to_excel(writer, sheet_name='Sensorinfo', index=False, header=['' for _ in new_df.columns])
                writer.save()

        # book = load_workbook(template_path_out)
        # writer = pd.ExcelWriter(template_path_out, engine='openpyxl')
        # writer.book = book

        # writer = pd.ExcelWriter(template_path_out, engine='openpyxl')

        # new_df.to_excel(writer, sheet_name='Sensorinfoo', index=False, header=['' for _ in new_df.columns])

        # writer.save()
        # writer.close()


        # m = DeliveryMetadataFile()
        # m.add_sensorinfo_from_file(r'C:\mw\temp_ctd_pre_system_data_root\cnv/sensorinfo.txt')
        # m.save_file(r'C:\mw\temp_ctd_pre_system_data_root\cnv/metadata.xlsx')




