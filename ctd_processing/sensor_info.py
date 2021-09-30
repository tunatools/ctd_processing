import pandas as pd
import openpyxl
import pathlib
import datetime

from ctd_processing import xmlcon
from ctd_processing import ctd_files
from ctd_processing import cnv


class InstrumentFile:
    def __init__(self, file_path):
        self._path = pathlib.Path(file_path)

        self._wb = openpyxl.load_workbook(self._path)
        self._sheets = self._wb.sheetnames

        self._info = {}

        self._save_info()

    def __str__(self):
        return f'InstrumentFile: {self._path}'

    def _save_info(self):
        self._info = {}
        df = pd.read_excel(self._path, sheet_name='Instrument.xls', engine='openpyxl', skiprows=[0])
        df = df.fillna('')
        for i in df.index:
            cnv_name = str(df.iloc[i]['CNV_NAME'])
            if cnv_name == '':
                continue
            self._info.setdefault(cnv_name, {})
            sensor_string = str(df.iloc[i]['SENSOR_ID'])
            if sensor_string is '':
                self._info[cnv_name]['all'] = df.iloc[i].to_dict()
            else:
                sensor_list = [item.strip() for item in sensor_string.split(',')]
                for sensor in sensor_list:
                    self._info[cnv_name][sensor] = df.iloc[i].to_dict()

    def get_info_for_parameter_and_sensor_id(self, parameter, sensor_id):
        """ Returns information from self._info. Matches key in self._info where key is part of "parameter" """
        for key, info in self._info.items():
            if key in parameter:
                return info.get(str(sensor_id))


class SensorInfoFile:
    def __init__(self, instrument_file: InstrumentFile):
        self.instrument_file = instrument_file
        self._stem = None
        self._save_path = None
        self._data = None

    def __str__(self):
        return f'SensorInfoFile: {self._stem}'

    def create_file_from_cnv_file(self, cnv_file_path):
        """
        Created a sensor info file with information in given cnv file.
        The sensor info file is created at the same location as the cnv file.
        """
        path = pathlib.Path(cnv_file_path)
        self._stem = path.stem
        self._save_path = pathlib.Path(path.parent, f'{self._stem}.sensorinfo')
        self._save_data_from_cnv(path)
        self._save_file()

    def _save_data_from_cnv(self, path):
        self._data = []
        cnv_info = xmlcon.CNVfileXML(path).get_sensor_info()
        channel_mapping = cnv.get_sensor_id_and_paramater_mapping_from_cnv(path)
        ctd_file_obj = ctd_files.get_ctd_files_object(path)

        instrument = path.name.split('_')[0]
        columns = [col for col in get_sensor_info_columns() if col not in ['VALIDFR', 'VALIDTO']]
        self._data.append('\t'.join(columns))
        for info in cnv_info:
            row_list = []
            instrument_info = self.instrument_file.get_info_for_parameter_and_sensor_id(parameter=info['parameter'],
                                                                                        sensor_id=info['serial_number'])
            if not instrument_info:
                print('NOT', info['serial_number'], info['parameter'])
                continue
            for col in columns:
                value = str(instrument_info.get(col, ''))
                if col == 'SENSOR_ID':
                    value = info['serial_number']
                elif col == 'CALIB_DATE':
                    value = info['calibration_date'].strftime('%Y-%m-%d')
                elif value:
                    if col == 'PARAMETER_REPORTED':
                        value = channel_mapping.get(info['serial_number'])
                    if info['parameter'][-1] == '2':
                        if col == 'PARAM_SIMPLE':
                            value = value + '2'
                        elif col == 'PARAM':
                            split_value = value.split('_')
                            split_value[-2] = split_value[-2] + '2'
                            value = '_'.join(split_value)
                else:
                    if col == 'INSTRUMENT_ID':
                        value = instrument + ctd_file_obj.instrument_number
                    elif col == 'INSTRUMENT_PROD':
                        if instrument.startswith('SBE'):
                            value = 'Seabird'
                    elif col == 'INSTRUMENT_MOD':
                        value = '911plus'
                    elif col == 'INSTRUMENT_SERIE':
                        value = ctd_file_obj.instrument_number
                    elif col == 'TIME':
                        value = ctd_file_obj.time.strftime('%Y-%m-%d')

                row_list.append(value)
            self._data.append('\t'.join(row_list))

    def _save_file(self):
        with open(self._save_path, 'w') as fid:
            fid.write('\n'.join(self._data))


class SensorInfoItem:
    """
    Holds information about a sensor INSTRUMENT_SERIE - PARAM_REPORTED kombination.
    """
    def __init__(self):
        self._all_columns = get_sensor_info_columns()
        self._columns = [col for col in self._all_columns if col not in ['VALIDFR', 'VALIDTO']]
        self._key = ()
        self._data = {}
        self._valid_from = None
        self._valid_to = None

    def _check_columns(self, data):
        if not all([col in self._columns for col in data.keys()]):
            raise Exception('Invalid data to SensorInfoItem')

    @staticmethod
    def get_key(data):
        return (data['INSTRUMENT_SERIE'], data['PARAM_REPORTED'])

    @staticmethod
    def _get_time_object(time_string):
        return datetime.datetime.strptime(time_string, '%Y-%m-%d')

    @staticmethod
    def _get_time_string(datetime_object):
        return datetime_object.strftime('%Y-%m-%d')

    @property
    def valid_from(self):
        return self._valid_from

    @property
    def valid_to(self):
        return self._valid_to

    def add_data(self, data):
        self._check_columns(data)
        if not self._key:
            self._add_first_data(data)
        else:
            return self._add_additional_data(data)
        return self._key

    def _add_first_data(self, data):
        self._key = self.get_key(data)
        self._data = data
        self._set_valid_from(data)
        self._set_valid_to(data)

    def _add_additional_data(self, data):
        if self.get_key(data) != self._key:
            return False
        self._set_valid_from(data)
        self._set_valid_to(data)
        return self._key

    def _set_valid_from(self, data):
        if not self._valid_from:
            self._valid_from = self._get_time_object(data['TIME'])
        else:
            self._valid_from = min(self._valid_from, self._get_time_object(data['TIME']))

    def _set_valid_to(self, data):
        if not self._valid_to:
            self._valid_to = self._get_time_object(data['TIME'])
        else:
            self._valid_to = max(self._valid_to, self._get_time_object(data['TIME']))

    def get_info(self):
        info = self._data.copy()
        info['VALIDFR'] = self._valid_from.strftime('%Y-%m-%d')
        info['VALIDTO'] = self._valid_to.strftime('%Y-%m-%d')
        return info


class SensorInfoFiles:

    def __init__(self, directory):
        self._directory = pathlib.Path(directory)
        self._paths = [path for path in self._directory.iterdir() if path.suffix == '.sensorinfo']
        self._sensor_info_items = {}
        self._save_info()

    def _save_info(self):
        self._sensor_info_items = {}
        for path in self._paths:
            header = None
            with open(path) as fid:
                for r, line in enumerate(fid):
                    split_line = [item.strip() for item in line.strip().split('\t')]
                    if r == 0:
                        header = split_line
                        continue
                    data = dict(zip(header, split_line))
                    key = SensorInfoItem.get_key(data)
                    obj = self._sensor_info_items.setdefault(key, SensorInfoItem())
                    obj.add_data(data)

    def write_to_file(self):
        path = pathlib.Path(self._directory, 'sensorinfo.txt')
        columns = get_sensor_info_columns()
        lines = []
        lines.append('\t'.join(columns))
        for key in sorted(self._sensor_info_items):
            info = self._sensor_info_items[key].get_info()
            lines.append('\t'.join([info.get(col, '') for col in columns]))

        with open(path, 'w') as fid:
            fid.write('\n'.join(lines))


def get_sensor_info_columns():
    path = pathlib.Path(pathlib.Path(__file__).parent, 'resources', 'sensor_info_columns.txt')
    columns = []
    with open(path) as fid:
        for line in fid:
            sline = line.strip()
            if not sline:
                continue
            columns.append(sline)
    return columns


def get_sensor_info_object(instrumentinfo_file):
    instrument_file = InstrumentFile(instrumentinfo_file)
    sensor_info = SensorInfoFile(instrument_file)
    return sensor_info


if __name__ == '__main__':
    cnv_file = r'C:\mw\temp_ctd_pre_system_data_root\cnv/SBE09_1387_20210413_1113_77SE_00_0278.cnv'
    sif = SensorInfoFiles(r'C:\mw\temp_ctd_pre_system_data_root\cnv')
    sif.write_to_file()
    instrument_file_path = r'C:\mw\git\ctd_config/instrumentinfo.xlsx'
    # # e = pd.read_excel(r'C:\mw\temp_svea\temp_filer/Instruments.xlsx', engine='openpyxl')
    # # e = openpyxl.load_workbook(file_path)
    # # df = pd.read_excel(file_path, sheet_name='SBE Dissolved Oxygen Sensors', engine='openpyxl')
    #
    # cnv_file = r'C:\mw\temp_ctd_pre_system_data_root\cnv/SBE09_1387_20210413_1113_77SE_00_0278.cnv'
    #
    instrument_file = InstrumentFile(instrument_file_path)
    #
    # sensor_info = SensorInfoFile(instrument_file)
    # sensor_info.create_file_from_cnv_file(cnv_file)
    #
    cnv_info = xmlcon.CNVfileXML(cnv_file).get_sensor_info()
